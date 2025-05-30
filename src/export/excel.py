# src/export/excel.py
# Standardized Excel output – now with Timeline, Impact & Quality sheets
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, List

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.data.cleaner import DataCleaner


class ExcelExporter:
    """Export analysis results to a multi-tab Excel workbook."""

    TEMPLATE_PATH = "templates/nexus_analysis_template.xlsx"  # still optional

    # Sheet names (legacy + new)
    SUMMARY_SHEET = "Nexus Summary"
    DETAILS_SHEET = "State Details"
    DATA_SHEET = "Source Data"
    TIMELINE_SHEET = "Nexus Timeline"
    IMPACT_SHEET = "Financial Impact"
    QUALITY_SHEET = "Data Quality"

    # ────────────────────────────────────────────────────────────────
    # PUBLIC API
    # ────────────────────────────────────────────────────────────────
    @staticmethod
    def export_results(
        results: List[Dict],
        sales_data: pd.DataFrame,
        output_path: str,
        client_name: str = "Client",
    ):
        results_df = pd.DataFrame(results)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # 1. Executive / summary sheet (keeps your original styling)
            ExcelExporter._create_summary(writer, results_df, client_name)

            # 2. Detailed results (original)
            ExcelExporter._create_state_details(writer, results_df)

            # 3. Timeline analysis (NEW)
            ExcelExporter._create_timeline(writer, results_df)

            # 4. Financial-impact estimates (NEW)
            ExcelExporter._create_financial_impact(writer, results_df, sales_data)

            # 5. Data-quality sheet (NEW)
            quality = DataCleaner.validate_data_quality(sales_data)
            ExcelExporter._create_quality_sheet(writer, quality)

            # 6. Raw-data sample (original)
            sales_data.head(1000).to_excel(
                writer, sheet_name=ExcelExporter.DATA_SHEET, index=False
            )

            # Global column auto-fit
            ExcelExporter._autosize_columns(writer.book)

    # ────────────────────────────────────────────────────────────────
    # SHEET BUILDERS
    # ────────────────────────────────────────────────────────────────
    @staticmethod
    def _create_summary(writer, results_df, client_name):
        stats = {
            "Client": client_name,
            "Analysis Date": datetime.now().strftime("%Y-%m-%d"),
            "Total States Analyzed": len(results_df),
            "States with Nexus": int(results_df["has_nexus"].sum()),
            "Earliest Breach": results_df.loc[
                results_df["has_nexus"], "breach_date"
            ].min()
            or "None",
        }
        df = pd.DataFrame(list(stats.items()), columns=["Metric", "Value"])
        df.to_excel(writer, sheet_name=ExcelExporter.SUMMARY_SHEET, index=False)

        ws = writer.sheets[ExcelExporter.SUMMARY_SHEET]
        # Preserve your header formatting
        ws.insert_rows(1)
        ws.merge_cells("A1:B1")
        ws["A1"] = f"Nexus Analysis Summary – {client_name}"
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

    @staticmethod
    def _create_state_details(writer, results_df):
        results_df.to_excel(
            writer, sheet_name=ExcelExporter.DETAILS_SHEET, index=False
        )
        ws = writer.sheets[ExcelExporter.DETAILS_SHEET]
        # Header formatting (legacy)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="366092")
        # Conditional formatting for nexus rows
        for row in range(2, len(results_df) + 2):
            if ws[f"B{row}"].value:  # has_nexus == True
                for col in ["A", "B", "C", "D", "E", "F"]:
                    ws[f"{col}{row}"].fill = PatternFill(
                        "solid", fgColor="FFE6E6"
                    )

    @staticmethod
    def _create_timeline(writer, results_df):
        nexus = results_df[results_df["has_nexus"]].copy()
        if nexus.empty:
            return
        nexus["breach_date"] = pd.to_datetime(nexus["breach_date"])
        nexus = nexus.sort_values("breach_date")
        nexus["Days Since Breach"] = (
            datetime.now() - nexus["breach_date"]
        ).dt.days
        cols = ["state", "breach_date", "breach_type", "Days Since Breach"]
        nexus[cols].to_excel(
            writer, sheet_name=ExcelExporter.TIMELINE_SHEET, index=False
        )

    @staticmethod
    def _create_financial_impact(writer, results_df, sales_data):
        rows = []
        for _, r in results_df[results_df["has_nexus"]].iterrows():
            state, breach_date = r["state"], pd.to_datetime(r["breach_date"])
            mask = (sales_data["state"] == state) & (
                pd.to_datetime(sales_data["date"]) >= breach_date
            )
            total = sales_data.loc[mask, "gross_sales"].sum()
            est_rate = 0.065  # simple avg; refine later
            liability = total * est_rate
            rows.append(
                {
                    "State": state,
                    "Breach Date": breach_date.strftime("%Y-%m-%d"),
                    "Sales Since Breach": f"${total:,.2f}",
                    "Est. Tax Rate": f"{est_rate:.1%}",
                    "Est. Liability": f"${liability:,.2f}",
                    "Penalty (10%)": f"${liability*0.10:,.2f}",
                }
            )
        if rows:
            pd.DataFrame(rows).to_excel(
                writer, sheet_name=ExcelExporter.IMPACT_SHEET, index=False
            )

    @staticmethod
    def _create_quality_sheet(writer, quality: Dict):
        df = pd.json_normalize(quality, sep=".").T.reset_index()
        df.columns = ["Metric", "Value"]
        df.to_excel(
            writer, sheet_name=ExcelExporter.QUALITY_SHEET, index=False
        )

    # ────────────────────────────────────────────────────────────────
    # UTILITIES
    # ────────────────────────────────────────────────────────────────
    @staticmethod
    def _autosize_columns(workbook: openpyxl.Workbook):
        """Simple auto-width for every sheet."""
        for ws in workbook.worksheets:
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col)
                max_len = max(
                    len(str(c.value)) if c.value is not None else 0
                    for c in ws[letter]
                )
                ws.column_dimensions[letter].width = min(max_len + 2, 50)
